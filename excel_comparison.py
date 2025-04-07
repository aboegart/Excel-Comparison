import argparse
import csv
import logging
import os

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

if os.path.exists('excel_comparison.log'):
    os.remove('excel_comparison.log')

logging.basicConfig(filename='excel_comparison.log', level=logging.INFO)


def find_client_nr_column(ws):
    for i, column in enumerate(ws.iter_cols(values_only=True)):
        if column[0].lower() == 'client nr':
            return i
    logging.error("'Client nr' column not found in worksheet.")
    raise KeyError("Could not find 'Client nr' column in worksheet.")


def get_client_nr_set(ws):
    client_nr_column = find_client_nr_column(ws)
    return set(column for column in list(
        ws.iter_cols(min_col=client_nr_column + 1, max_col=client_nr_column + 1, min_row=2, values_only=True))[0])


def adjust_column_width(ws):
    for i, column in enumerate(ws.iter_cols(), start=1):
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except Exception:
                pass
        adjusted_width = (max_length * 1.2)
        ws.column_dimensions[get_column_letter(i)].width = adjusted_width


def main(file1, file2, q, copy_only_new_clients=False, case_sensitive=False):
    headers_to_copy = ['Client nr', 'Client', 'Address', 'NIP']

    try:
        if not os.path.exists(file1) or not os.path.exists(file2):
            logging.error(f"Files {file1} or {file2} do not exist.")
            q.put(("update_label", "Error: Both input files must exist."))
            return

        q.put(("update_label", "Starting comparison..."))
        q.put(("update_progress", 0))

        _, file1_ext = os.path.splitext(file1)
        _, file2_ext = os.path.splitext(file2)

        if file1_ext == '.csv':
            with open(file1, 'r') as f:
                reader = csv.reader(f)
                headers = next(reader)
                data1 = list(reader)
        else:
            wb1 = load_workbook(file1)
            wb2 = load_workbook(file2)
            ws1 = wb1.active
            ws2 = wb2.active
            headers = [cell.value.lower() if cell.value is not None else '' for cell in ws1[1]]
            data1 = list(ws1.iter_rows(min_row=2, values_only=True))

        logging.info("Starting comparison...\n")

        client_nrs_ws1 = get_client_nr_set(ws1)
        client_nrs_ws2 = get_client_nr_set(ws2)

        new_clients = client_nrs_ws2 - client_nrs_ws1
        q.put(("update_progress", 50))
        q.put(("update_label", "Halfway done..."))

        original_headers = [cell.value if cell.value is not None else '' for cell in ws2[1]]
        headers = [header.lower() for header in original_headers]
        new_workbook = Workbook()
        new_sheet = new_workbook.active
        new_sheet.append([original_headers[headers.index(header.lower())] for header in headers_to_copy])

        clients_in_new_workbook = set()

        if case_sensitive:
            headers = [cell.value if cell.value is not None else '' for cell in ws2[1]]
        else:
            headers = [cell.value.lower() if cell.value is not None else '' for cell in ws2[1]]

        if not copy_only_new_clients:
            for row in ws1.iter_rows(min_row=2, values_only=True):
                client_nr = row[headers.index('client nr')]
                if client_nr not in clients_in_new_workbook:
                    new_sheet.append([row[headers.index(header.lower())] for header in headers_to_copy])
                    clients_in_new_workbook.add(client_nr)

        for row in ws2.iter_rows(min_row=2, values_only=True):
            client_nr = row[headers.index('client nr')]
            if client_nr in new_clients and client_nr not in clients_in_new_workbook:
                new_sheet.append([row[headers.index(header.lower())] for header in headers_to_copy])
                clients_in_new_workbook.add(client_nr)

        adjust_column_width(new_sheet)
        data = [[cell for cell in row] for row in new_sheet.iter_rows(values_only=True)]
        logging.info("\nComparison finished!")
        q.put(("update_progress", 100))
        q.put(("update_label", "Comparison finished!"))
        return data

    except Exception as e:
        logging.error(str(e))
        q.put(("update_label", "Error: " + str(e)))
        return
    finally:
        q.put(("update_button", "normal"))


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Compare two Excel files.')
    parser.add_argument('file1', type=str, help='First Excel file')
    parser.add_argument('file2', type=str, help='Second Excel file')
    args = parser.parse_args()

    main(args.file1, args.file2)
